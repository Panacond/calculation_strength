# алгоритм по созданию таблицы нагрузок
# определение суммирования идет через пересчет порядкового номера заголовка
# Описание нагрузки     плотность   толщина     ширина      длина       коэф
# функции
# просто читает все файлы csv и записывает в exel
import os
import fnmatch
import csv
# Запись файлов Excel
import openpyxl
# загрузка стилей
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
# воспомгательные к общему файлу
def read_load(l):
    l = l[1:]
    # создание вложенного списка
    for i in range(len(l)):
        while l[i][-1] =='':
            del l[i][-1]
        if len(l[i])>2:
            while l[i][-2] =='':
                del l[i][-2]
    for i in range(len(l)):
        x = l[i]
        for j in range(len(x)):
            if j !=0:
                x[j]=x[j].replace(',','.')
                x[j]=float(x[j])
        if len(x) == 6:
            x += ['кг']
        if len(x) == 5:
            x += ['кг/м']
        if len(x) == 4:
            x += ['кг/м.кв']
        l[i]=x
    return l

# функция открытия и запись в список с добавлением описания
def read_file(a):
    a = str(a) + '.csv'
    l=[]
    with open(a,'r',newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            l = l + [row]
    l = read_load(l)
    return l
#запись в таблицу
def write_csv(a,MT):
    a = str(a)[:-1] + '1' + '.csv'
    with open(a,'w',newline='') as file:
        writer = csv.writer(file)
        writer.writerows(MT)

def sort_dict(a):
    # словарь нагрузок
    # сортировка видов нагрузок в обратном направлении от величины
    # возвращает список отсортированных ключей
    kj = []
    b = a.values()
    b = str(b)[13:-2].split(', ')
    b = [float(i) for i in b]
    b.sort()
    k1 = a.keys()
    k = str(k1)[12:-3].split('\', \'')
    for i in b:
        for j in k:
            if i == a[j]:
                kj.append(j)
    return kj

# Функция создания списка с суммами
def table_load(l):
    # создание таблицы в форме списка
    MT=[['Вид нагрузки','Ед. изм.','Нормативная нагрузка',
         'Коэффициент надежности по нагрузке','Расчетная нагрузка'],
        ['1','2','3','4','5']]
    sum_place = [0,0] #сумма в одном загружений
    sum_local = [0,0] #сумма по этажу
    sum_all=[0,0] #общая сумма всех нагрузок
    list_name = [] #список названий (постоянные, временные...)
    dict_load_n = {} #словарь загружений нормативных
    dict_load_r = {} #словарь загружений расчетных
    i_type = '' # текущий тип нагрузки(постоянная)
    for ii in range(len(l)):
        i = l[ii]
        # создание заголовка
        if len(i) == 1:
            i_type = i[0] # Задание типа нагрузки для добавления сумм в словарь
            if ii < len(l) and len(l[ii+1]) != 1 and not i[0] in list_name:#  выбор названий загружений
                list_name.append(i[0])
                dict_load_n[i[0]] = 0
                dict_load_r[i[0]] = 0
            # перед заголовком
            # добавление сумм
            if not i[0] in list_name and ii > 1:
                MT +=[['Итого',index[1],sum_local[0],'',sum_local[1]]]
                MT +=[['Всего',index[1],sum_place[0],'',sum_place[1]]]
                sum_place = [0,0]
                sum_local = [0,0]
            if   i[0] in list_name and len(l[ii-1]) != 1:
                MT +=[['Итого',index[1],sum_local[0],'',sum_local[1]]]
                sum_local = [0,0]
                # сам заголовок
            MT +=[i]
        # основная строка
        else:
            dense = i[1]/(1000)
            if i[3] == 1000:
                text = i[0]+' G='+str(round(i[2]))+' кг/м.кв'
            elif i[3]==1:
                i[3]=1000
                text = i[0]+' G='+str(round(i[2]))+' кг'
            else:
                text = i[0]+' G='+str(round(i[2]))+' кг/м.куб t='+str(round(i[3]))+' мм'
            a = i[2:-1]
            for j in a:
                dense *= j
            index = [
                text ,i[-1],round(dense/i[1]),i[1],round(dense)]
            MT +=[index]
            # print(i)
            # print(index)
            # суммирование
            sum_local =[index[2]+sum_local[0],index[-1]+sum_local[-1]]
            sum_all=[index[2]+sum_all[0],index[-1]+sum_all[-1]]
            sum_place = [index[2]+sum_place[0],index[-1]+sum_place[-1]]
            dict_load_n[i_type] += index[2]
            dict_load_r[i_type] += index[-1]
            # добавление в конце таблицы итоговых результатов
            if ii + 1 == len(l):
                MT +=[['Итого',index[1],sum_local[0],'',sum_local[1]]]
                MT +=[['Всего',index[1],sum_place[0],'',sum_place[1]]]
                for i in sort_dict(dict_load_n):
                    MT +=[['Всего ' + i.lower() ,index[1],dict_load_n[i],'',dict_load_r[i]]]
                MT +=[['Сумма всех нагрузок',index[1],sum_all[0],'',sum_all[1]]]
                MT +=[['Фактическая постоянная расчетная распределенная нагрузка от веса конструкций\nсоставляет '+str(sum_all[1])+' '+index[1]+' (нормативная нагрузка '+ str(sum_all[0])+' '+index[1]+ ')']]
                
    len_dop = len(list_name)
    return MT, len_dop

# все в итоговую функцию
def calc(MT, name):
    # создаем новый excel-файл
    wb = openpyxl.Workbook()
    # добавляем новый лист
    wb.create_sheet(title = 'Сбор нагрузок', index = 0)
    # получаем лист, с которым будем работать
    sheet = wb['Сбор нагрузок']
    # оформление
    #!!!! создаем именованный стиль !!!!
    ns = NamedStyle(name='row1')
    ns.font = Font(name = 'Times New Roman', size=12)
    # Выравнивание текста
    ns.alignment = Alignment(horizontal='center', vertical='center',
                                wrapText='wrapText')
    # заливка
    n = '00C0C0C0'
    ns.fill = PatternFill(fill_type='solid', start_color= n, end_color= n )
    # Рамка
    border = Side(style='medium', color='00000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns)
    # второй стиль
    ns2 = NamedStyle(name='row2')
    ns2.font = Font(name = 'Times New Roman', size=12, italic= 'italic')
    # Выравнивание текста
    ns2.alignment = Alignment(horizontal='center', vertical='center')
    # Рамка
    border = Side(style='medium', color='00000000')
    ns2.border = Border(left=border, top=border, right=border, bottom=border)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns2)
    # Жирный стиль подписей выровненный по центру
    ns3 = NamedStyle(name='row3')
    ns3.font = Font(name = 'Times New Roman', size=12, bold='+')
    # Выравнивание текста
    ns3.alignment = Alignment(horizontal='center', vertical=None,
                 textRotation=0, wrapText=None, shrinkToFit=None, indent=0, relativeIndent=0,
                 justifyLastLine=None, readingOrder=0, text_rotation=None,
                 wrap_text=None, shrink_to_fit=None, mergeCell=None)
    # Рамка
    border = Side(style='medium', color='00000000')
    ns3.border = Border(left=border, top=border, right=border, bottom=border)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns3)
    # Жирный стиль подписей остальных
    ns4 = NamedStyle(name='row4')
    ns4.font = Font(name = 'Times New Roman', size=12, bold='+')
    # Рамка
    border = Side(style='medium', color='00000000')
    ns4.border = Border(left=border, top=border, right=border, bottom=border)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns4)
    # Колонки со 2 по 5
    ns5 = NamedStyle(name='row5')
    ns5.font = Font(name = 'Times New Roman', size=12)
    # Выравнивание текста
    ns5.alignment = Alignment(horizontal='center')
    # Рамка
    border = Side(style='medium', color='00000000')
    ns5.border = Border(left=border, top=border, right=border, bottom=border)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns5)
    # остальные 1-e столбцы
    ns6 = NamedStyle(name='row6')
    ns6.font = Font(name = 'Times New Roman', size=12)
    # Рамка
    border = Side(style='medium', color='00000000')
    ns6.border = Border(left=border, top=border, right=border, bottom=border)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns6)

    # последняя строка
    ns7 = NamedStyle(name='row7')
    ns7.font = Font(name = 'Times New Roman', size=12)
    # Рамка
    border = Side(style='medium', color='00000000')
    ns7.border = Border(left=border, top=border, right=border, bottom=border)
    ns7.alignment=Alignment(wrap_text=True, shrink_to_fit=False,indent=0)
    # вновь созданный именованный стиль надо зарегистрировать
    # для дальнейшего использования
    wb.add_named_style(ns7)


    #запись значений в строку
    for row in range(len(MT)):
        for col in range(len(MT[row])):
            value = str(MT[row][col])
            cell = sheet.cell(row = row + 1, column = col + 1)
            cell.value = value
        # назначение стилей
        if row == 0:
            # 1 строка шапка
            for col in range(0,5):
                sheet.cell(row = row + 1, column = col + 1).style = 'row1'
        elif row == 1:
            # 2 строка номера шапок
            for col in range(0,5):
                sheet.cell(row = row + 1, column = col + 1).style = 'row2'
        elif  row != len(MT) -1 and (row == 2  or (len(MT[row]) == 1 and (len(MT[row + 1]) == 1) )) :
            # Жирный стиль подписей выровненный по центру
            for col in range(0,5):
                sheet.cell(row = row + 1, column = col + 1).style = 'row3'
        elif row == len(MT) -1:
            # формат последней строки
            for col in range(0,1):
                sheet.cell(row = row + 1, column = col + 1).style = 'row7'
                sheet.merge_cells(range_string = None , start_row = row + 1, start_column = col + 1 , end_row = row + 2 , end_column = col + 5)
        elif (len(MT[row]) == 1) or MT[row][3] =='':
            # Жирный стиль остальных подписей
            for col in range(0,1):
                sheet.cell(row = row + 1, column = col + 1).style = 'row4'
            for col in range(1,5):
                # Колонки со 2 по 5
                sheet.cell(row = row + 1, column = col + 1).style = 'row5'
        else:
            for col in range(0,1):
                sheet.cell(row = row + 1, column = col + 1).style = 'row6'
            for col in range(1,5):
                sheet.cell(row = row + 1, column = col + 1).style = 'row5'

    # Настройка строк и столбцов
    sheet.row_dimensions[1].height = 99
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 15
    wb.save(name + '.xlsx')
    # Печать итогов
    print(name)
    for x in MT[-3:-1]:
        print(x)
        # вывод нужных значений нагрузок
    Q1 = MT[-3:-1][0][-3]
    Q2 = MT[-3:-1][0][-1]
    Q3 = MT[-2:][0][-3]
    Q4 = MT[-2:][0][-1]
    Q = [Q1, Q2, Q3, Q4]
    return Q, MT
# дополнение к существующему расчету
def pre_calc(name):
    l = read_file(name)
    MT, len_dop=table_load(l)
    return name, MT
def main():
    # отбор нужных файлов
    f = []# создание списка файлов и чтение из текущей папки списка файлов
    for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*load.csv'):
            f += [file]
    need_f = []# создание списка нужных файлов
    for i in f:# отбор нужных файлов
         if i[-4:] == '.csv':
            need_f += [i[:-4]]
    # создание таблиц
    for i in need_f:
        name, MT = pre_calc(i)
        calc(MT = MT, name = name)
    pass

if __name__ == '__main__':
    main()

# запуск из общего файла
def load_donload(l, name_file):
    name = name_file[:-4]
    l = read_load(l)
    MT = table_load(l)[0]
    Q, MT = calc(MT = MT, name = name)
    return Q, MT
